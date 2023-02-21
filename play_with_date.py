from datetime import datetime
import openpyxl
from openpyxl.styles import NamedStyle
from openpyxl.styles import Alignment

wb = openpyxl.load_workbook('american_date.xlsx')
sheet = wb['List1']
date_style = NamedStyle(name='datetime', number_format='DD/MM/YYYY')

for i in range(1, sheet.max_row + 1):
    my_date = datetime.strptime(sheet.cell(row=i, column=1).value, '%m/%d/%Y')
    europe_date = datetime.strftime(my_date, '%d/%m/%Y')
    sheet.cell(row=i, column=1).value = europe_date
    sheet.cell(row=i, column=1).style = date_style
    sheet.cell(row=i, column=1).alignment = Alignment(horizontal='right')
# TODO: Zamień stringi na daty.
wb.save('american_date1.xlsx')
