import openpyxl

wb = openpyxl.load_workbook('vegetables.xlsx')
sheet = wb['List1']

new_wb = openpyxl.Workbook()
new_sheet = new_wb.create_sheet(title='changed', index=0)

for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        new_sheet.cell(column, row).value = sheet.cell(row, column).value


new_wb.save('vegetable_transpose.xlsx')
